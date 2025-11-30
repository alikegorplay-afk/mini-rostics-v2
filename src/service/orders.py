import asyncio

from concurrent.futures import ThreadPoolExecutor

from sqlalchemy.ext.asyncio import async_sessionmaker, AsyncSession
from openpyxl.styles import Font, Alignment

from loguru import logger
import pandas as pd
import openpyxl

from ..schemas.order import OrderSchema
from ..core.const import REPORT_PATH
from ..managers.order_manager import OrderManager
from ..managers.product_manager import ProductManager

class OrderProductService:
    def __init__(self, session_maker: async_sessionmaker[AsyncSession]):
        self.order_manager = OrderManager(session_maker)
        self.product_manager = ProductManager(session_maker)
        self.executer = ThreadPoolExecutor(max_workers = 2)
        
    async def get_order_sum(self, order_id: int) -> float | None:
        order = await self.order_manager.get_order(order_id)
        if not order:
            return None
        return await self._get_order_sum(order)
    
    async def get_revenue(self) -> float:
        summa: float = 0
        for order in await self.order_manager.get_all_orders():
            if order.status == "Оплачен":
                order_sum = await self.get_order_sum(order.id)
                if order_sum is None:
                    logger.warning(f"Не удалось найти заказ по ID {order.id}")
                    continue
                summa += order_sum
        return round(summa, 2)
    
    async def generate_report(self) -> bool:
        try:
            summa = 0
            data = []
            for order in await self.order_manager.get_all_orders():
                data.append(
                    {
                        "ID": order.id,
                        "Сумма заказа": f"{await self._get_order_sum(order)} ₽",
                        "Статус заказа": order.status
                    }
                )
                if order.status == "Оплачен":
                    order_sum = await self.get_order_sum(order.id)
                    if order_sum is None:
                        logger.warning(f"Не удалось найти заказ по ID {order.id}")
                        continue
                    summa += order_sum
                    
            df = pd.DataFrame(data).style.apply(lambda x: ['text-align: center' for _ in x], subset=['ID'])          
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(
                executor = self.executer,
                func = lambda: self._create_xlsx(df, round(summa, 2))
            )
            return True
        except Exception as e:
            logger.error(f"Ошибка при получении репорта {e}")
            return False
        
    async def _get_order_sum(self, order: OrderSchema):
        ids = {x.product_id: x.count for x in order.items}
        items = await self.product_manager.get_products([x for x in ids.keys()])
        return round(sum([x.price * ids[x.id] for x in items]), 2)
    
    def _create_xlsx(self, df: pd.DataFrame, revenue: float):
        df.to_excel(REPORT_PATH, index=False)
        wb = openpyxl.load_workbook(REPORT_PATH)
        ws = wb.active
        
        for row in (ws[f'A{x}'] for x in range(1, ws.max_row + 1)):
            row.alignment = Alignment(horizontal='center')
        
        last_row = ws.max_row + 2
        ws[f'A{last_row}'] = f"Выручка: {revenue} ₽"
        
        ws.merge_cells(f'A{last_row}:C{last_row}')
        ws[f'A{last_row}'].font = Font(bold=True, size=12, color="00FF00")
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')
        
        wb.save(REPORT_PATH)